#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
מערכת המרת מחירונים - Price Converter
ממיר קבצי .gz של מחירונים לאקסל מעוצב
"""

import os
import gzip
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

def extract_gz_file(gz_path):
    """פותח קובץ gz ומחלץ את תוכן ה-XML"""
    try:
        # Try as gzip first
        with gzip.open(gz_path, 'rt', encoding='utf-8') as f:
            return f.read()
    except gzip.BadGzipFile:
        # If not gzipped, try as regular XML
        try:
            with open(gz_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            print(f"❌ שגיאה בקריאת הקובץ: {e}")
            return None
    except Exception as e:
        print(f"❌ שגיאה בפתיחת הקובץ: {e}")
        return None

def parse_xml_to_excel(xml_content, output_path):
    """ממיר XML לאקסל מעוצב"""
    try:
        # Parse XML
        root = ET.fromstring(xml_content)
        
        # Get store info
        chain_id = root.find('ChainId')
        chain_id = chain_id.text if chain_id is not None else 'לא ידוע'
        
        store_id = root.find('StoreId')
        store_id = store_id.text if store_id is not None else 'לא ידוע'
        
        sub_chain_id = root.find('SubChainId')
        sub_chain_id = sub_chain_id.text if sub_chain_id is not None else ''
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "מחירון מוצרים"
        
        # RTL support
        ws.sheet_view.rightToLeft = True
        
        # Headers
        headers = [
            'קוד מוצר', 
            'שם המוצר', 
            'יצרן', 
            'מחיר (₪)', 
            'יחידת מידה', 
            'כמות', 
            'מחיר ליחידה', 
            'תאריך עדכון', 
            'ארץ ייצור', 
            'הנחה אפשרית'
        ]
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Add data
        item_count = 0
        for item in root.findall('.//Item'):
            item_code = item.find('ItemCode')
            item_name = item.find('ItemNm')
            manufacturer = item.find('ManufacturerName')
            price = item.find('ItemPrice')
            unit_measure = item.find('UnitOfMeasure')
            quantity = item.find('Quantity')
            unit_price = item.find('UnitOfMeasurePrice')
            update_date = item.find('PriceUpdateDate')
            country = item.find('ManufactureCountry')
            allow_discount = item.find('AllowDiscount')
            
            row_data = [
                item_code.text if item_code is not None else '',
                item_name.text if item_name is not None else '',
                manufacturer.text if manufacturer is not None else '',
                price.text if price is not None else '',
                unit_measure.text if unit_measure is not None else '',
                quantity.text if quantity is not None else '',
                unit_price.text if unit_price is not None else '',
                update_date.text if update_date is not None else '',
                country.text if country is not None else '',
                'כן' if allow_discount is not None and allow_discount.text == '1' else 'לא'
            ]
            ws.append(row_data)
            item_count += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 3, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Align all cells to right
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Add info sheet
        info_ws = wb.create_sheet("מידע על המחירון")
        info_ws.sheet_view.rightToLeft = True
        
        info_data = [
            ['מזהה רשת:', chain_id],
            ['מזהה רשת משנה:', sub_chain_id],
            ['מזהה סניף:', store_id],
            ['מספר מוצרים:', str(item_count)],
            ['תאריך יצוא:', datetime.now().strftime('%d/%m/%Y %H:%M')],
            ['', ''],
            ['הופק על ידי:', 'מערכת המרת מחירונים - Price Converter']
        ]
        
        for row in info_data:
            info_ws.append(row)
        
        # Style info sheet
        for cell in info_ws['A']:
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='right')
        
        for cell in info_ws['B']:
            cell.alignment = Alignment(horizontal='right')
        
        info_ws.column_dimensions['A'].width = 20
        info_ws.column_dimensions['B'].width = 30
        
        # Save
        wb.save(output_path)
        
        return item_count, chain_id, store_id
        
    except Exception as e:
        print(f"❌ שגיאה בעיבוד הנתונים: {e}")
        import traceback
        traceback.print_exc()
        return None, None, None

def process_file(input_file):
    """מעבד קובץ אחד"""
    print(f"\n📂 מעבד קובץ: {os.path.basename(input_file)}")
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"❌ הקובץ לא נמצא: {input_file}")
        return False
    
    # Extract gz
    print("📦 פותח קובץ דחוס...")
    xml_content = extract_gz_file(input_file)
    
    if xml_content is None:
        return False
    
    # Create output filename
    base_name = os.path.basename(input_file)
    output_name = base_name.replace('.gz', '').replace('.xml', '') + '_מחירון.xlsx'
    output_path = os.path.join(os.path.dirname(input_file), output_name)
    
    # Convert to Excel
    print("📊 ממיר לאקסל...")
    item_count, chain_id, store_id = parse_xml_to_excel(xml_content, output_path)
    
    if item_count is None:
        return False
    
    # Success!
    print(f"\n✅ הצלחה!")
    print(f"📊 מספר מוצרים: {item_count}")
    print(f"🏪 רשת: {chain_id}, סניף: {store_id}")
    print(f"💾 נשמר ב: {output_path}")
    
    return True

def process_folder(folder_path):
    """מעבד את כל קבצי ה-gz בתיקייה"""
    gz_files = [f for f in os.listdir(folder_path) if f.endswith('.gz')]
    
    if not gz_files:
        print("❌ לא נמצאו קבצי .gz בתיקייה")
        return
    
    print(f"\n🔍 נמצאו {len(gz_files)} קבצים")
    
    success_count = 0
    for gz_file in gz_files:
        full_path = os.path.join(folder_path, gz_file)
        if process_file(full_path):
            success_count += 1
    
    print(f"\n✅ סיום! עובדו {success_count}/{len(gz_files)} קבצים בהצלחה")

def main():
    """תפריט ראשי"""
    print("=" * 60)
    print("🛒 מערכת המרת מחירונים - Price Converter")
    print("=" * 60)
    print()
    
    if len(sys.argv) > 1:
        # Run with command line argument
        input_path = sys.argv[1]
        if os.path.isfile(input_path):
            process_file(input_path)
        elif os.path.isdir(input_path):
            process_folder(input_path)
        else:
            print(f"❌ הנתיב לא חוקי: {input_path}")
    else:
        # Interactive mode
        print("בחר אפשרות:")
        print("1. המר קובץ בודד")
        print("2. המר את כל הקבצים בתיקייה")
        print("3. יציאה")
        print()
        
        choice = input("הקלד מספר (1/2/3): ").strip()
        
        if choice == '1':
            file_path = input("\nגרור את הקובץ לכאן (או הדבק נתיב): ").strip().strip('"')
            if os.path.exists(file_path):
                process_file(file_path)
            else:
                print("❌ הקובץ לא נמצא")
        
        elif choice == '2':
            folder_path = input("\nהדבק נתיב לתיקייה: ").strip().strip('"')
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                process_folder(folder_path)
            else:
                print("❌ התיקייה לא נמצאה")
        
        elif choice == '3':
            print("👋 להתראות!")
            return
        else:
            print("❌ בחירה לא חוקית")
    
    print("\n" + "=" * 60)
    input("לחץ Enter לסיום...")

if __name__ == "__main__":
    main()